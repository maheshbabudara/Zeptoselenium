import openpyxl as p
from time import sleep

def reader(file_name,sheet_name):
    workbook=p.load_workbook(file_name)
    sheet=workbook[sheet_name]
    rows=sheet.max_row
    cols=sheet.max_column
    credentails=[]
    for r in range(2,rows+1):
        cred=[]
        for c in range(1,cols+1):
            data=sheet.cell(r,c).value
            cred.append(data)
        credentails.append(cred)
    return credentails
